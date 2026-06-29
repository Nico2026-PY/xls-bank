import os
import re
import sys
import json
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Versión v0.2.15: preparación comercial, licencia propietaria y pantalla Acerca de XlsBank.

APP_NAME = "XlsBank"
APP_VERSION = "v0.2.15"
APP_AUTHOR = "Nicolás Mellado"
APP_COPYRIGHT = "Copyright © 2026 Nicolás Mellado. Todos los derechos reservados."

# ============================================================
# PANTALLA DE CARGA ANIMADA
# ============================================================

def ruta_recurso(ruta_relativa):
    """
    Obtiene la ruta correcta al ejecutar desde Python o desde un .exe.

    Soporta:
    - Desarrollo: proyecto/assets
    - Desarrollo desde app_src/app.py
    - PyInstaller: carpeta temporal _MEIPASS
    """
    ruta_relativa = Path(ruta_relativa)

    if hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS) / ruta_relativa

    archivo_actual = Path(__file__).resolve()
    carpeta_app_src = archivo_actual.parent
    carpeta_proyecto = carpeta_app_src.parent

    posibles = [
        carpeta_proyecto / ruta_relativa,
        carpeta_app_src / ruta_relativa,
        Path.cwd() / ruta_relativa,
    ]

    for ruta in posibles:
        if ruta.exists():
            return ruta

    return carpeta_proyecto / ruta_relativa


class PantallaCarga:
    """Muestra el GIF, una barra azul y el porcentaje antes de abrir la app."""

    def __init__(self, root, ruta_gif, al_terminar, ancho=760, alto=720):
        self.root = root
        self.ruta_gif = Path(ruta_gif)
        self.al_terminar = al_terminar
        self.ancho = ancho
        self.alto = alto
        self.indice = 0
        self.frames = []
        self.duraciones = []

        self.color_fondo = '#0B1220'
        self.color_barra = '#2F7CF6'

        self.ventana = tk.Toplevel(root)
        self.ventana.overrideredirect(True)
        self.ventana.attributes('-topmost', True)
        self.ventana.configure(bg=self.color_fondo)
        self.ventana.geometry(self._geometria_centrada())

        contenedor = tk.Frame(self.ventana, bg=self.color_fondo)
        contenedor.pack(fill='both', expand=True)

        self.label = tk.Label(
            contenedor,
            bg=self.color_fondo,
            borderwidth=0,
            highlightthickness=0
        )
        self.label.pack(pady=(35, 18))

        estilo = ttk.Style(self.ventana)
        estilo.theme_use('clam')
        estilo.configure(
            'Splash.Horizontal.TProgressbar',
            troughcolor='#162033',
            background=self.color_barra,
            bordercolor='#162033',
            lightcolor=self.color_barra,
            darkcolor=self.color_barra,
            thickness=18
        )

        marco_barra = tk.Frame(contenedor, bg=self.color_fondo)
        marco_barra.pack(fill='x', padx=100)

        self.barra = ttk.Progressbar(
            marco_barra,
            orient='horizontal',
            mode='determinate',
            maximum=100,
            value=0,
            style='Splash.Horizontal.TProgressbar'
        )
        self.barra.pack(fill='x')

        self.porcentaje = tk.Label(
            contenedor,
            text='0%',
            bg=self.color_fondo,
            fg='#1D4ED8',
            font=('Segoe UI', 12, 'bold')
        )
        self.porcentaje.pack(pady=(10, 0))

        self._cargar_frames()
        self._animar()

    def _geometria_centrada(self):
        self.root.update_idletasks()
        ancho_pantalla = self.root.winfo_screenwidth()
        alto_pantalla = self.root.winfo_screenheight()
        x = (ancho_pantalla - self.ancho) // 2
        y = (alto_pantalla - self.alto) // 2
        return f'{self.ancho}x{self.alto}+{x}+{y}'

    def _cargar_frames(self):
        if not self.ruta_gif.exists():
            raise FileNotFoundError(f'No se encontró el GIF:\n{self.ruta_gif}')

        gif = Image.open(self.ruta_gif)
        cantidad = getattr(gif, 'n_frames', 1)

        for numero in range(cantidad):
            gif.seek(numero)

            # Se conserva la transparencia del GIF y luego se compone
            # sobre el mismo fondo oscuro de la ventana.
            frame = gif.convert('RGBA').copy()
            frame.thumbnail((520, 520), Image.Resampling.LANCZOS)

            fondo = Image.new(
                'RGBA',
                (560, 540),
                self.color_fondo
            )

            x = (fondo.width - frame.width) // 2
            y = (fondo.height - frame.height) // 2
            fondo.alpha_composite(frame, (x, y))

            self.frames.append(
                ImageTk.PhotoImage(fondo.convert('RGB'))
            )
            self.duraciones.append(
                max(30, int(gif.info.get('duration', 70)))
            )

        gif.close()

        if not self.frames:
            raise RuntimeError('El GIF no contiene cuadros para reproducir.')

    def _animar(self):
        self.label.configure(image=self.frames[self.indice])

        total = max(1, len(self.frames) - 1)
        progreso = round((self.indice / total) * 100)

        self.barra['value'] = progreso
        self.porcentaje.configure(text=f'{progreso}%')

        demora = self.duraciones[self.indice]

        if self.indice < len(self.frames) - 1:
            self.indice += 1
            self.ventana.after(demora, self._animar)
        else:
            self.barra['value'] = 100
            self.porcentaje.configure(text='100%')
            self.ventana.after(350, self._cerrar)

    def _cerrar(self):
        self.ventana.destroy()
        self.al_terminar()


# ============================================================
# PROCESADOR BANCARIO POR EMPRESA
# Genera un Excel con una hoja por empresa y bloques por banco/cuenta.
# Soporta estructura BPN, Galicia y Patagonia como los ejemplos enviados.
# ============================================================

BANCOS_CLAVES = {
    'BPN': ['bpn', 'provincia del neuquen', 'neuquen'],
    'GALICIA': ['galicia'],
    'PATAGONIA': ['patagonia'],
    'MERCADO_PAGO': ['mercado pago', 'mercadopago', 'mp '],
    'NACION': ['nacion', 'bna'],
    'MACRO': ['macro'],
    'SANTANDER': ['santander'],
    'BBVA': ['bbva', 'frances'],
    'ICBC': ['icbc'],
}

# Las empresas NO se dejan escritas en el código para poder publicar
# el repositorio sin exponer nombres reales, CUITs ni datos internos.
# Se cargan desde un archivo privado externo: empresas_config.json
CONFIG_EMPRESAS_ARCHIVO = 'empresas_config.json'
EMPRESAS_CLAVES = {}

COLUMNAS_SALIDA = [
    'Cuenta', 'Fecha Valor', 'Fecha Operación', 'Movimiento Fecha-Valor',
    'Descripción', 'Detalle', 'Importe', 'Saldo', 'Categoria', 'Referencia', 'Etiquetas'
]


def norm(txt):
    txt = '' if txt is None else str(txt)
    txt = txt.strip().lower()
    txt = txt.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ü','u')
    return re.sub(r'\s+', ' ', txt)


def rutas_posibles_config_empresas():
    """
    Devuelve ubicaciones posibles del archivo privado de empresas.

    No se incluye este archivo en GitHub ni en el ZIP de release.
    Orden recomendado:
    1) Variable de entorno BANCOS_EMPRESAS_CONFIG.
    2) Carpeta del usuario en Windows: %APPDATA%\\Procesador_Bancario\\empresas_config.json.
    3) Carpeta del .exe o del código, útil para pruebas locales.
    """
    rutas = []

    ruta_env = os.environ.get('BANCOS_EMPRESAS_CONFIG', '').strip()
    if ruta_env:
        rutas.append(Path(ruta_env))

    for variable in ['APPDATA', 'LOCALAPPDATA']:
        base = os.environ.get(variable, '').strip()
        if base:
            rutas.append(Path(base) / 'Procesador_Bancario' / CONFIG_EMPRESAS_ARCHIVO)

    try:
        rutas.append(Path(sys.executable).resolve().parent / CONFIG_EMPRESAS_ARCHIVO)
    except Exception:
        pass

    try:
        rutas.append(Path(__file__).resolve().parent / CONFIG_EMPRESAS_ARCHIVO)
    except Exception:
        pass

    rutas.append(Path.cwd() / CONFIG_EMPRESAS_ARCHIVO)

    # Quita repetidos conservando orden.
    vistas = []
    unicas = []
    for ruta in rutas:
        clave = str(ruta).lower()
        if clave not in vistas:
            vistas.append(clave)
            unicas.append(ruta)
    return unicas


def normalizar_config_empresas(data):
    """
    Acepta estos formatos privados:

    Formato recomendado:
    {
      "empresas": {
        "EMPRESA_1": ["clave archivo", "razon social"],
        "EMPRESA_2": {"claves": ["otra clave"], "cuit": "opcional"}
      }
    }

    También acepta lista:
    {
      "empresas": [
        {"nombre": "EMPRESA_1", "claves": ["clave"], "cuit": "opcional"}
      ]
    }

    El CUIT puede estar en el archivo privado para futuras reglas, pero esta
    versión solo usa nombre y claves para detectar empresa.
    """
    if not isinstance(data, dict):
        return {}

    empresas = data.get('empresas', data)
    salida = {}

    if isinstance(empresas, list):
        for item in empresas:
            if not isinstance(item, dict):
                continue
            nombre = str(item.get('nombre', '')).strip().upper()
            claves = item.get('claves', [])
            if isinstance(claves, str):
                claves = [claves]
            claves = [str(c).strip() for c in claves if str(c).strip()]
            if nombre and claves:
                salida[nombre] = claves
        return salida

    if isinstance(empresas, dict):
        for nombre, valor in empresas.items():
            nombre_limpio = str(nombre).strip().upper()
            claves = []
            if isinstance(valor, dict):
                claves = valor.get('claves', [])
            elif isinstance(valor, str):
                claves = [valor]
            elif isinstance(valor, list):
                claves = valor

            claves = [str(c).strip() for c in claves if str(c).strip()]
            if nombre_limpio and claves:
                salida[nombre_limpio] = claves

    return salida


def cargar_empresas_claves():
    """Carga palabras clave privadas de empresas desde empresas_config.json."""
    for ruta in rutas_posibles_config_empresas():
        try:
            if not ruta.exists():
                continue
            data = json.loads(ruta.read_text(encoding='utf-8-sig'))
            empresas = normalizar_config_empresas(data)
            if empresas:
                return empresas
        except Exception:
            # No se detiene la app por una configuración privada dañada.
            continue

    # Sin config privada, la app sigue funcionando por fallback:
    # usa la primera palabra del archivo como empresa y muestra advertencia.
    return {}


EMPRESAS_CLAVES = cargar_empresas_claves()


def limpiar_numero(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return 0.0
    s = s.replace('$', '').replace(' ', '')
    # Formato argentino: 1.234.567,89
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except Exception:
        return 0.0


def limpiar_fecha(valor):
    """
    Convierte fechas de bancos a fecha real.

    Corrige especialmente Patagonia, que a veces trae la fecha como texto
    tipo 14/6/26 o 14/05/26. Si se ordena eso como texto, Excel/Python
    mezcla meses. Con esta función siempre se ordena como fecha real.
    """
    if valor is None:
        return pd.NaT

    try:
        if pd.isna(valor):
            return pd.NaT
    except Exception:
        pass

    if isinstance(valor, pd.Timestamp):
        return valor.to_pydatetime()

    if isinstance(valor, datetime):
        return valor

    # Por si viene como date de Python, sin ser datetime.
    if hasattr(valor, 'year') and hasattr(valor, 'month') and hasattr(valor, 'day') and not isinstance(valor, str):
        try:
            return datetime(valor.year, valor.month, valor.day)
        except Exception:
            pass

    # Fechas seriales de Excel.
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        try:
            numero = float(valor)
            if 20000 <= numero <= 80000:
                fecha = pd.to_datetime('1899-12-30') + pd.to_timedelta(numero, unit='D')
                return fecha.to_pydatetime()
        except Exception:
            pass

    s = str(valor).strip()
    if s == '' or s.lower() in ['nan', 'none', 'nat']:
        return pd.NaT

    # Argentina: día/mes/año. Soporta 14/6/26, 14/06/2026, 14-06-26, etc.
    fecha = pd.to_datetime(s, errors='coerce', dayfirst=True)

    if pd.isna(fecha):
        return pd.NaT

    return fecha.to_pydatetime()


def normalizar_fechas_movimientos(df):
    """Normaliza columnas de fecha para que todas se ordenen igual."""
    df = df.copy()
    for col in ['Fecha Valor', 'Fecha Operación']:
        if col in df.columns:
            df[col] = df[col].apply(limpiar_fecha)
    return df


def ordenar_por_fecha_real(df, descendente=True):
    """Ordena movimientos por fecha real, no por texto."""
    df = normalizar_fechas_movimientos(df)

    if 'Fecha Valor' not in df.columns:
        return df

    fecha_orden = pd.to_datetime(df['Fecha Valor'], errors='coerce')

    if 'Fecha Operación' in df.columns:
        fecha_operacion = pd.to_datetime(df['Fecha Operación'], errors='coerce')
        fecha_orden = fecha_orden.fillna(fecha_operacion)

    df['_FechaOrden'] = fecha_orden
    df = df.sort_values(
        '_FechaOrden',
        ascending=not descendente,
        na_position='last',
        kind='mergesort'
    )
    return df.drop(columns=['_FechaOrden'])

def obtener_ultimo_saldo_actualizado(df):
    """
    Devuelve el saldo de la fecha más actualizada del bloque.

    Antes se ordenaba de menor a mayor y se tomaba .last(), pero cuando
    había varios movimientos el mismo día podía agarrar un saldo intermedio
    de esa fecha. Ahora toma el primer saldo visible al ordenar por fecha
    descendente, respetando el orden original del extracto para empates
    del mismo día.
    """
    if df is None or len(df) == 0 or 'Saldo' not in df.columns:
        return 0

    tmp = normalizar_fechas_movimientos(df).copy()
    tmp['_OrdenOriginal'] = range(len(tmp))

    if 'Fecha Valor' in tmp.columns:
        fecha_orden = pd.to_datetime(tmp['Fecha Valor'], errors='coerce')
    else:
        fecha_orden = pd.Series(pd.NaT, index=tmp.index)

    if 'Fecha Operación' in tmp.columns:
        fecha_operacion = pd.to_datetime(tmp['Fecha Operación'], errors='coerce')
        fecha_orden = fecha_orden.fillna(fecha_operacion)

    tmp['_FechaOrden'] = fecha_orden

    con_fecha = tmp[tmp['_FechaOrden'].notna()].copy()
    if con_fecha.empty:
        saldos = tmp['Saldo'].dropna()
        return saldos.iloc[0] if len(saldos) else 0

    con_fecha = con_fecha.sort_values(
        ['_FechaOrden', '_OrdenOriginal'],
        ascending=[False, True],
        na_position='last',
        kind='mergesort'
    )

    saldos = con_fecha['Saldo'].dropna()
    return saldos.iloc[0] if len(saldos) else 0


def crear_resumen_general(todos):
    """Arma el resumen general usando el saldo real de la fecha más reciente."""
    claves = ['Empresa', 'Banco', 'Tipo Cuenta', 'Bloque']
    filas = []

    for valores, df_bloque in todos.groupby(claves, dropna=False, sort=False):
        if not isinstance(valores, tuple):
            valores = (valores,)

        fila = dict(zip(claves, valores))
        importes = pd.to_numeric(df_bloque['Importe'], errors='coerce').fillna(0)

        fila['Cant. Movimientos'] = int(importes.count())
        fila['Total Ingresos'] = importes[importes > 0].sum()
        fila['Total Egresos'] = importes[importes < 0].sum()
        fila['Neto'] = importes.sum()
        fila['Último Saldo'] = obtener_ultimo_saldo_actualizado(df_bloque)

        filas.append(fila)

    return pd.DataFrame(
        filas,
        columns=claves + [
            'Cant. Movimientos',
            'Total Ingresos',
            'Total Egresos',
            'Neto',
            'Último Saldo'
        ]
    )



def detectar_banco_detalle(path, df_raw=None):
    """Devuelve banco detectado y método de detección."""
    n = norm(Path(path).stem)
    for banco, claves in BANCOS_CLAVES.items():
        if any(c in n for c in claves):
            return banco, 'nombre_archivo'

    if df_raw is not None:
        try:
            texto = ' '.join(df_raw.head(30).astype(str).fillna('').values.ravel())
            t = norm(texto)
            for banco, claves in BANCOS_CLAVES.items():
                if any(c in t for c in claves):
                    return banco, 'contenido_archivo'
        except Exception:
            pass

    return 'BANCO', 'desconocido'


def detectar_banco(path, df_raw=None):
    banco, _ = detectar_banco_detalle(path, df_raw)
    return banco


def detectar_empresa_detalle(path, df_raw=None):
    """Devuelve empresa detectada y método de detección."""
    n = norm(Path(path).stem)
    for emp, claves in EMPRESAS_CLAVES.items():
        if any(c in n for c in claves):
            return emp, 'nombre_archivo'

    # Si no encuentra por nombre, busca titularidad dentro de las primeras filas.
    if df_raw is not None:
        try:
            texto = ' '.join(df_raw.head(25).astype(str).fillna('').values.ravel())
            t = norm(texto)
            for emp, claves in EMPRESAS_CLAVES.items():
                if any(c in t for c in claves):
                    return emp, 'contenido_archivo'
        except Exception:
            pass

    # Último recurso: primera palabra del archivo.
    fallback = Path(path).stem.split()[0].upper()
    return fallback, 'fallback'


def detectar_empresa(path, df_raw=None):
    empresa, _ = detectar_empresa_detalle(path, df_raw)
    return empresa


def detectar_tipo_cuenta(cuenta, archivo):
    txt = norm(str(cuenta) + ' ' + Path(archivo).stem)
    if 'caja' in txt or ' ca ' in f' {txt} ' or txt.startswith('ca') or 'ca ' in txt:
        return 'Caja de Ahorro'
    if 'corriente' in txt or 'cta cte' in txt or 'cc ' in txt or txt.startswith('cc'):
        return 'Cuenta Corriente'
    if 'mercado pago' in txt or 'mercadopago' in txt or ' mp ' in f' {txt} ':
        return 'Billetera Virtual'
    return 'Cuenta'


def read_excel_any(path, header=None):
    """
    Lee .xlsx, .xls y .csv.
    Para CSV usa separador ; porque Mercado Pago exporta así.
    """
    ext = Path(path).suffix.lower()

    if ext == '.csv':
        try:
            return pd.read_csv(
                path,
                header=header,
                dtype=object,
                sep=';',
                encoding='utf-8-sig'
            )
        except UnicodeDecodeError:
            return pd.read_csv(
                path,
                header=header,
                dtype=object,
                sep=';',
                encoding='latin1'
            )

    if ext == '.xls':
        return pd.read_excel(path, header=header, dtype=object, engine='xlrd')

    return pd.read_excel(path, header=header, dtype=object, engine='openpyxl')

def encontrar_fila_encabezado(df_raw, nombres):
    objetivos = [norm(x) for x in nombres]
    for i in range(len(df_raw)):
        vals = [norm(x) for x in df_raw.iloc[i].tolist()]
        aciertos = sum(1 for o in objetivos if o in vals)
        if aciertos >= max(2, min(3, len(objetivos))):
            return i
    return None


def procesar_bpn(path):
    df = read_excel_any(path, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    cuenta = df['Cuenta'].dropna().iloc[0] if 'Cuenta' in df.columns and not df['Cuenta'].dropna().empty else ''
    out = pd.DataFrame()
    out['Cuenta'] = df.get('Cuenta', cuenta)
    out['Fecha Valor'] = df.get('Fecha Valor', '')
    out['Fecha Operación'] = df.get('Fecha Operación', '')
    out['Movimiento Fecha-Valor'] = df.get('Movimiento Fecha-Valor', '')
    out['Descripción'] = df.get('Descripción', '')
    out['Detalle'] = df.get('Detalle', '')
    out['Importe'] = df['Importe'].apply(limpiar_numero) if 'Importe' in df.columns else ''
    out['Saldo'] = df['Saldo'].apply(limpiar_numero) if 'Saldo' in df.columns else ''
    out['Categoria'] = df.get('Categoria', '')
    out['Referencia'] = df.get('Referencia', '')
    out['Etiquetas'] = df.get('Etiquetas', '')
    return out[COLUMNAS_SALIDA]


def procesar_galicia(path):
    df = read_excel_any(path, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    cuenta = 'GALICIA'
    deb = df['Débitos'].apply(limpiar_numero) if 'Débitos' in df.columns else 0
    cre = df['Créditos'].apply(limpiar_numero) if 'Créditos' in df.columns else 0
    out = pd.DataFrame()
    out['Cuenta'] = cuenta
    out['Fecha Valor'] = df.get('Fecha', '')
    out['Fecha Operación'] = df.get('Fecha', '')
    out['Movimiento Fecha-Valor'] = ''
    out['Descripción'] = df.get('Descripción', '')
    # Detalle: une observación + cliente + leyendas si existen
    partes = []
    for c in ['Observaciones Cliente', 'Leyendas Adicionales 1', 'Leyendas Adicionales 2', 'Leyendas Adicionales 3']:
        if c in df.columns:
            partes.append(df[c].fillna('').astype(str))
    if partes:
        detalle = partes[0]
        for p in partes[1:]: detalle = detalle.str.cat(p, sep=' ', na_rep='')
    else:
        detalle = ''
    out['Detalle'] = detalle
    out['Importe'] = cre - deb
    out['Saldo'] = df['Saldo'].apply(limpiar_numero) if 'Saldo' in df.columns else 0
    out['Categoria'] = ''  # Galicia no trae una columna real llamada Categoría; no se inventa.
    out['Referencia'] = df.get('Número de Comprobante', '')
    out['Etiquetas'] = ''
    return out[COLUMNAS_SALIDA]


def limpiar_celda_texto(valor):
    """Convierte una celda a texto limpio, evitando nan/none/nat."""
    if valor is None:
        return ''

    try:
        if pd.isna(valor):
            return ''
    except Exception:
        pass

    texto = str(valor).strip()
    if texto.lower() in ['nan', 'none', 'nat']:
        return ''
    return re.sub(r'\s+', ' ', texto)


def es_codigo_cuenta_patagonia(texto):
    """Detecta cuentas Patagonia tipo CC$ 385-123012668-000 o CA$ 385-..."""
    if not texto:
        return False

    t = limpiar_celda_texto(texto)
    tn = norm(t)

    # Evita devolver etiquetas del encabezado.
    etiquetas = ['cuenta:', 'cuenta', 'titularidad:', 'titularidad', 'movimientos de cuenta']
    if tn in etiquetas:
        return False

    patrones = [
        r'\b(?:cc|ca|cc\$|ca\$)\s*\$?\s*\d{2,4}[- ]\d{5,12}[- ]\d{1,4}\b',
        r'\b\d{2,4}[- ]\d{5,12}[- ]\d{1,4}\b',
    ]
    return any(re.search(p, tn, flags=re.IGNORECASE) for p in patrones)


def extraer_valor_despues_de_etiqueta(raw, etiqueta, filas_busqueda=25, columnas_busqueda=10):
    """
    Busca una etiqueta en las primeras filas y devuelve el valor cercano.

    Banco Patagonia suele traer:
    Cuenta:      | CC$ 385-123012668-000
    Titularidad: | EMPRESA EJEMPLO

    En algunos Excel el valor queda en la misma celda, en la celda de al lado
    o en una fila cercana. Esta función cubre esos casos sin depender de una
    columna fija.
    """
    etiqueta_norm = norm(etiqueta).replace(':', '')
    max_filas = min(filas_busqueda, len(raw))
    max_cols = min(columnas_busqueda, raw.shape[1])

    for i in range(max_filas):
        for j in range(max_cols):
            celda = limpiar_celda_texto(raw.iat[i, j])
            if not celda:
                continue

            celda_norm = norm(celda).replace(':', '')
            if etiqueta_norm not in celda_norm:
                continue

            # Caso 1: "Cuenta: CC$ 385-..." todo en la misma celda.
            partes = re.split(r':', celda, maxsplit=1)
            if len(partes) == 2 and limpiar_celda_texto(partes[1]):
                valor = limpiar_celda_texto(partes[1])
                if norm(valor) not in [etiqueta_norm, 'cuenta', 'titularidad']:
                    return valor

            # Caso 2: valor en celdas cercanas hacia la derecha.
            for jj in range(j + 1, min(j + 6, raw.shape[1])):
                valor = limpiar_celda_texto(raw.iat[i, jj])
                if valor and norm(valor).replace(':', '') not in [etiqueta_norm, 'cuenta', 'titularidad']:
                    return valor

            # Caso 3: valor en la fila siguiente o subsiguiente.
            for ii in range(i + 1, min(i + 4, len(raw))):
                for jj in range(j, min(j + 6, raw.shape[1])):
                    valor = limpiar_celda_texto(raw.iat[ii, jj])
                    if valor and norm(valor).replace(':', '') not in [etiqueta_norm, 'cuenta', 'titularidad']:
                        return valor

    return ''


def extraer_cuenta_patagonia(raw):
    """Devuelve número de cuenta y titularidad de extractos Banco Patagonia."""
    cuenta = extraer_valor_despues_de_etiqueta(raw, 'Cuenta:')
    titularidad = extraer_valor_despues_de_etiqueta(raw, 'Titularidad:')

    # Si la búsqueda por etiqueta encontró una celda incorrecta, hace un barrido
    # por patrón de cuenta dentro del encabezado.
    if not es_codigo_cuenta_patagonia(cuenta):
        cuenta_detectada = ''
        max_filas = min(25, len(raw))
        max_cols = min(12, raw.shape[1])
        for i in range(max_filas):
            for j in range(max_cols):
                valor = limpiar_celda_texto(raw.iat[i, j])
                if es_codigo_cuenta_patagonia(valor):
                    cuenta_detectada = valor
                    break
            if cuenta_detectada:
                break
        cuenta = cuenta_detectada or cuenta

    # Limpieza final por si quedó texto extra al lado.
    if cuenta:
        m = re.search(
            r'((?:CC|CA)\s*\$?\s*\d{2,4}[- ]\d{5,12}[- ]\d{1,4}|\d{2,4}[- ]\d{5,12}[- ]\d{1,4})',
            cuenta,
            flags=re.IGNORECASE
        )
        if m:
            cuenta = m.group(1).strip()

    return cuenta, titularidad



def extraer_textos_binarios_xls(path):
    """
    Extrae cadenas legibles de archivos .XLS antiguos.

    Algunos extractos de Banco Patagonia guardan el número de cuenta como
    texto dentro del archivo, pero pandas/xlrd puede devolver la plantilla
    o celdas vacías. Este fallback revisa el binario para encontrar valores
    reales como CC$ 385-123012668-000.
    """
    textos = []
    try:
        data = Path(path).read_bytes()
    except Exception:
        return textos

    # Cadenas ANSI/CP1252 típicas de .xls BIFF.
    patron = re.compile(rb'[\x20-\x7E\x80-\xFF]{3,}')
    for match in patron.finditer(data):
        try:
            s = match.group(0).decode('cp1252', errors='ignore')
        except Exception:
            continue
        s = limpiar_celda_texto(s)
        if s and s not in textos:
            textos.append(s)

    # También intenta UTF-16LE por si el .xls guarda strings Unicode.
    try:
        texto_wide = data.decode('utf-16le', errors='ignore')
        for s in re.findall(r'[A-Za-zÁÉÍÓÚÜÑáéíóúüñ0-9$.,:/() _\-]{3,}', texto_wide):
            s = limpiar_celda_texto(s)
            if s and s not in textos:
                textos.append(s)
    except Exception:
        pass

    return textos


def extraer_cuenta_patagonia_desde_archivo(path):
    """Fallback para leer cuenta/titularidad Patagonia desde el binario .XLS."""
    textos = extraer_textos_binarios_xls(path)
    cuenta = ''
    titularidad = ''

    patron_cuenta = re.compile(
        r'((?:CC|CA)\s*\$?\s*\d{2,4}[- ]\d{5,12}[- ]\d{1,4}|\d{2,4}[- ]\d{5,12}[- ]\d{1,4})',
        flags=re.IGNORECASE
    )

    indice_cuenta = None
    for idx, texto in enumerate(textos):
        m = patron_cuenta.search(texto)
        if m:
            cuenta = m.group(1).strip().rstrip('#').strip()
            indice_cuenta = idx
            break

    if indice_cuenta is not None:
        # Suele venir justo después del número de cuenta.
        for texto in textos[indice_cuenta + 1: indice_cuenta + 8]:
            tn = norm(texto)
            if not texto or es_codigo_cuenta_patagonia(texto):
                continue
            if any(x in tn for x in ['cuenta', 'titularidad', 'movimientos de cuenta', 'banco patagonia']):
                continue
            if re.search(r'[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]', texto) and len(texto) >= 3:
                titularidad = texto.strip(' #\t')
                break

    return cuenta, titularidad

def procesar_patagonia(path):
    raw = read_excel_any(path, header=None)
    fila = encontrar_fila_encabezado(raw, ['Fecha', 'Descripción', 'Débito', 'Crédito', 'Saldo'])
    if fila is None:
        raise ValueError('No se encontró encabezado Patagonia con Fecha/Descripción/Débito/Crédito/Saldo')
    headers = raw.iloc[fila].tolist()
    data = raw.iloc[fila+1:].copy()
    data.columns = [str(h).strip() if str(h) != 'nan' else '' for h in headers]
    data = data.dropna(how='all')
    data = data[data.get('Fecha').notna()] if 'Fecha' in data.columns else data

    cuenta, titularidad = extraer_cuenta_patagonia(raw)

    # Fallback especial para Patagonia .XLS: en algunos archivos xlrd lee
    # placeholders o celdas vacías, aunque el número real está guardado en el binario.
    if not es_codigo_cuenta_patagonia(cuenta):
        cuenta_bin, titular_bin = extraer_cuenta_patagonia_desde_archivo(path)
        if cuenta_bin:
            cuenta = cuenta_bin
        if not titularidad and titular_bin:
            titularidad = titular_bin

    deb = data['Débito'].apply(limpiar_numero) if 'Débito' in data.columns else 0
    cre = data['Crédito'].apply(limpiar_numero) if 'Crédito' in data.columns else 0
    out = pd.DataFrame()
    out['Cuenta'] = cuenta or 'PATAGONIA'
    out['Fecha Valor'] = data.get('Fecha', '')
    out['Fecha Operación'] = data.get('Fecha', '')
    out['Movimiento Fecha-Valor'] = ''
    out['Descripción'] = data.get('Descripción', '')
    out['Detalle'] = data.get('Referencia', '')
    out['Importe'] = cre - deb
    out['Saldo'] = data['Saldo'].apply(limpiar_numero) if 'Saldo' in data.columns else 0
    out['Categoria'] = ''  # Patagonia no trae una categoría real; se deja vacío para no repetir la titularidad.
    out['Referencia'] = data.get('Referencia', '')
    out['Etiquetas'] = ''
    return out[COLUMNAS_SALIDA]



def col_mp(df, nombre, default=''):
    """Devuelve columna de Mercado Pago si existe, si no devuelve default."""
    if nombre in df.columns:
        return df[nombre]
    return default


def describir_tipo_mp(transaction_type):
    """Traduce tipos de operación de Mercado Pago a textos más claros."""
    t = norm(transaction_type)

    traducciones = {
        'settlement': 'Liquidación',
        'payment': 'Pago',
        'refund': 'Devolución',
        'chargeback': 'Contracargo',
        'withdrawal': 'Retiro de dinero',
        'payout': 'Retiro de dinero',
        'transfer': 'Transferencia',
    }

    return traducciones.get(t, str(transaction_type).strip() or 'Movimiento')


def describir_medio_mp(payment_method_type):
    """Traduce medios de pago de Mercado Pago."""
    t = norm(payment_method_type)

    traducciones = {
        'available_money': 'Dinero disponible',
        'account_money': 'Dinero en cuenta',
        'debit_card': 'Tarjeta débito',
        'credit_card': 'Tarjeta crédito',
        'bank_transfer': 'Transferencia bancaria',
        'ticket': 'Efectivo / cupón',
        'digital_currency': 'Moneda digital',
    }

    return traducciones.get(t, str(payment_method_type).strip() or '')


def describir_subunidad_mp(sub_unit):
    """Traduce subunidad/canal de Mercado Pago."""
    t = norm(sub_unit)

    traducciones = {
        'qr': 'QR',
        'point': 'Point',
        'checkout': 'Checkout',
        'available_money': 'Dinero disponible',
        'debit_card': 'Tarjeta débito',
        'credit_card': 'Tarjeta crédito',
        'bank_transfer': 'Transferencia bancaria',
    }

    return traducciones.get(t, str(sub_unit).strip() or '')

def procesar_mercado_pago(path):
    """
    Procesa reportes de Mercado Pago Argentina.

    Soporta:
    1) Formato viejo/en español.
    2) CSV nuevo con columnas SOURCE_ID, TRANSACTION_TYPE, REAL_AMOUNT, etc.
    """
    df = read_excel_any(path, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    columnas = set(df.columns)

    # ============================================================
    # FORMATO NUEVO CSV MERCADO PAGO
    # ============================================================
    columnas_mp_csv = {
        'SOURCE_ID',
        'PAYMENT_METHOD_TYPE',
        'TRANSACTION_TYPE',
        'TRANSACTION_AMOUNT',
        'TRANSACTION_DATE',
        'REAL_AMOUNT'
    }

    if columnas_mp_csv.issubset(columnas):
        out = pd.DataFrame(index=df.index)

        fecha_operacion = pd.to_datetime(
            df['TRANSACTION_DATE'],
            errors='coerce',
            utc=True
        ).dt.tz_convert(None)

        if 'SETTLEMENT_DATE' in df.columns:
            fecha_valor = pd.to_datetime(
                df['SETTLEMENT_DATE'],
                errors='coerce',
                utc=True
            ).dt.tz_convert(None)
        elif 'MONEY_RELEASE_DATE' in df.columns:
            fecha_valor = pd.to_datetime(
                df['MONEY_RELEASE_DATE'],
                errors='coerce',
                utc=True
            ).dt.tz_convert(None)
        else:
            fecha_valor = fecha_operacion

        importe_neto = df['REAL_AMOUNT'].apply(limpiar_numero)

        tipo_raw = df['TRANSACTION_TYPE'].fillna('').astype(str)
        medio_raw = df['PAYMENT_METHOD_TYPE'].fillna('').astype(str)

        tipo_claro = tipo_raw.apply(describir_tipo_mp)
        medio_claro = medio_raw.apply(describir_medio_mp)

        if 'SUB_UNIT' in df.columns:
            sub_raw = df['SUB_UNIT'].fillna('').astype(str)
            sub_claro = sub_raw.apply(describir_subunidad_mp)
        else:
            sub_raw = pd.Series('', index=df.index)
            sub_claro = pd.Series('', index=df.index)

        descripcion = tipo_claro

        for extra in [medio_claro, sub_claro]:
            extra = extra.fillna('').astype(str).str.strip()
            descripcion = descripcion.where(
                extra.eq(''),
                descripcion + ' - ' + extra
            )

        detalle_partes = []

        for c in ['BUSINESS_UNIT', 'SUB_UNIT', 'PAYMENT_METHOD_TYPE']:
            if c in df.columns:
                detalle_partes.append(df[c].fillna('').astype(str))

        if detalle_partes:
            detalle = detalle_partes[0]
            for parte in detalle_partes[1:]:
                detalle = detalle.str.cat(parte, sep=' | ', na_rep='')
        else:
            detalle = pd.Series('', index=df.index)

        out['Cuenta'] = 'Mercado Pago Argentina'
        out['Fecha Valor'] = fecha_valor
        out['Fecha Operación'] = fecha_operacion
        out['Movimiento Fecha-Valor'] = ''
        out['Descripción'] = descripcion
        out['Detalle'] = detalle
        out['Importe'] = importe_neto

        # Mercado Pago no informa saldo de cuenta en este CSV.
        out['Saldo'] = ''

        out['Categoria'] = ''
        out['Referencia'] = df['SOURCE_ID'].fillna('').astype(str)
        out['Etiquetas'] = ''

        return out[COLUMNAS_SALIDA]

    # ============================================================
    # FORMATO VIEJO / ESPAÑOL MERCADO PAGO
    # ============================================================
    columnas_necesarias = [
        'FECHA DE APROBACIÓN',
        'TIPO DE OPERACIÓN',
        'MONTO NETO DE LA OPERACIÓN QUE IMPACTÓ TU DINERO'
    ]

    faltantes = [c for c in columnas_necesarias if c not in df.columns]
    if faltantes:
        raise ValueError(
            'No se reconoce formato Mercado Pago. Faltan columnas: '
            + ', '.join(faltantes)
        )

    out = pd.DataFrame(index=df.index)

    fecha_aprobacion = pd.to_datetime(
        df['FECHA DE APROBACIÓN'],
        errors='coerce',
        utc=True
    ).dt.tz_convert(None)

    fecha_liquidacion = (
        pd.to_datetime(
            df['FECHA DE LIQUIDACIÓN DEL DINERO'],
            errors='coerce',
            utc=True
        ).dt.tz_convert(None)
        if 'FECHA DE LIQUIDACIÓN DEL DINERO' in df.columns
        else fecha_aprobacion
    )

    importe_neto = df['MONTO NETO DE LA OPERACIÓN QUE IMPACTÓ TU DINERO'].apply(limpiar_numero)

    tipo_operacion = df['TIPO DE OPERACIÓN'].fillna('').astype(str)

    if 'DETALLE DE LA VENTA' in df.columns:
        detalle_venta = df['DETALLE DE LA VENTA'].fillna('').astype(str)
        descripcion = tipo_operacion + ' - ' + detalle_venta
    else:
        descripcion = tipo_operacion

    partes = []
    for c in ['PAGADOR', 'MEDIO DE PAGO', 'BANCO DE ORIGEN', 'NOMBRE DE LOCAL', 'CANAL DE VENTA']:
        if c in df.columns:
            partes.append(df[c].fillna('').astype(str))

    if partes:
        detalle = partes[0]
        for p in partes[1:]:
            detalle = detalle.str.cat(p, sep=' | ', na_rep='')
    else:
        detalle = pd.Series('', index=df.index)

    if 'ID DE OPERACIÓN EN MERCADO PAGO' in df.columns:
        referencia = df['ID DE OPERACIÓN EN MERCADO PAGO'].fillna('').astype(str)
    elif 'NÚMERO DE IDENTIFICACIÓN' in df.columns:
        referencia = df['NÚMERO DE IDENTIFICACIÓN'].fillna('').astype(str)
    else:
        referencia = pd.Series('', index=df.index)

    out['Cuenta'] = 'Mercado Pago Argentina'
    out['Fecha Valor'] = fecha_liquidacion
    out['Fecha Operación'] = fecha_aprobacion
    out['Movimiento Fecha-Valor'] = ''
    out['Descripción'] = descripcion
    out['Detalle'] = detalle
    out['Importe'] = importe_neto
    out['Saldo'] = ''
    out['Categoria'] = ''
    out['Referencia'] = referencia
    out['Etiquetas'] = ''

    return out[COLUMNAS_SALIDA]


def asegurar_salida_conservadora(mov, banco):
    """
    Evita que el Excel final muestre datos inventados o reinterpretados.

    Regla general:
    - No completar Categoría con empresa, titularidad, canal, plataforma ni tipo
      salvo que el banco traiga una columna real de Categoría/Categoria.
    - No completar Saldo con 0 si el banco no informa saldo.
    - No alterar archivos originales: la app solo lee los Excel de entrada y crea
      un nuevo Excel de salida.
    """
    mov = mov.copy()

    if 'Categoria' in mov.columns and banco in ['GALICIA', 'PATAGONIA', 'MERCADO_PAGO']:
        mov['Categoria'] = ''

    if 'Etiquetas' in mov.columns and banco in ['GALICIA', 'PATAGONIA', 'MERCADO_PAGO']:
        mov['Etiquetas'] = ''

    if 'Saldo' in mov.columns and banco == 'MERCADO_PAGO':
        mov['Saldo'] = ''

    return mov

def procesar_archivo(path):
    raw_preview = None
    try:
        raw_preview = read_excel_any(path, header=None)
    except Exception:
        raw_preview = None

    banco = detectar_banco(path, raw_preview)
    empresa = detectar_empresa(path, raw_preview)

    if banco == 'BPN':
        mov = procesar_bpn(path)
    elif banco == 'GALICIA':
        mov = procesar_galicia(path)
    elif banco == 'PATAGONIA':
        mov = procesar_patagonia(path)
    elif banco == 'MERCADO_PAGO':
        mov = procesar_mercado_pago(path)
    else:
        # Intenta como BPN/tabular estándar.
        # Si falla, la vista previa lo mostrará como Error.
        mov = procesar_bpn(path)

    mov = asegurar_salida_conservadora(mov, banco)

    # IMPORTANTE: normaliza fechas antes de juntar bancos/empresas.
    # Evita que Patagonia ordene fechas como texto: 14/6/26 antes que 14/05/26.
    mov = normalizar_fechas_movimientos(mov)

    cuenta = str(mov['Cuenta'].dropna().iloc[0]) if not mov['Cuenta'].dropna().empty else banco
    tipo = detectar_tipo_cuenta(cuenta, path)
    banco_titulo = banco.replace('_', ' ').title()
    etiqueta = f'{banco_titulo} – {tipo} {empresa.title()}'
    mov['Empresa'] = empresa
    mov['Banco'] = banco
    mov['Tipo Cuenta'] = tipo
    mov['Bloque'] = etiqueta
    mov['Archivo Origen'] = Path(path).name
    return mov


def crear_advertencia(tipo, archivo, banco, empresa, mensaje, solucion):
    return {
        'Tipo': tipo,
        'Archivo': archivo,
        'Banco': banco,
        'Empresa': empresa,
        'Mensaje': mensaje,
        'Solución sugerida': solucion,
    }


def obtener_fecha_orden_serie(df):
    if df is None or len(df) == 0:
        return pd.Series(dtype='datetime64[ns]')

    if 'Fecha Valor' in df.columns:
        fecha_orden = pd.to_datetime(df['Fecha Valor'], errors='coerce')
    else:
        fecha_orden = pd.Series(pd.NaT, index=df.index)

    if 'Fecha Operación' in df.columns:
        fecha_operacion = pd.to_datetime(df['Fecha Operación'], errors='coerce')
        fecha_orden = fecha_orden.fillna(fecha_operacion)

    return fecha_orden


def obtener_rango_fechas(df):
    fechas = obtener_fecha_orden_serie(df).dropna()
    if fechas.empty:
        return None, None
    return fechas.min().to_pydatetime(), fechas.max().to_pydatetime()


def validar_movimientos(path, banco, empresa, tipo_cuenta, bloque, mov, metodo_banco='desconocido', metodo_empresa='fallback'):
    """Genera advertencias preventivas para mostrar en la vista previa y en el Excel."""
    archivo = Path(path).name
    advertencias = []

    if banco == 'BANCO' or metodo_banco == 'desconocido':
        advertencias.append(crear_advertencia(
            'Advertencia', archivo, banco, empresa,
            'No se pudo reconocer el banco por nombre ni por contenido.',
            'Renombrar el archivo incluyendo el banco o revisar si el formato corresponde a un banco soportado.'
        ))

    if metodo_empresa == 'fallback':
        advertencias.append(crear_advertencia(
            'Advertencia', archivo, banco, empresa,
            'La empresa no fue detectada por palabras clave; se usó la primera palabra del archivo.',
            'Renombrar el archivo con una palabra clave configurada o cargar/editar el archivo privado empresas_config.json.'
        ))

    if mov is None or len(mov) == 0:
        advertencias.append(crear_advertencia(
            'Error', archivo, banco, empresa,
            'El archivo no tiene movimientos procesables.',
            'Revisar si el Excel fue exportado completo y si contiene la tabla de movimientos.'
        ))
        return advertencias

    fechas = obtener_fecha_orden_serie(mov)
    fechas_invalidas = int(fechas.isna().sum())
    if fechas_invalidas > 0:
        advertencias.append(crear_advertencia(
            'Advertencia', archivo, banco, empresa,
            f'Hay {fechas_invalidas} movimiento(s) con fecha vacía o no reconocida.',
            'Revisar la columna de fecha en el extracto antes de usar el resumen final.'
        ))

    if 'Saldo' in mov.columns and banco != 'MERCADO_PAGO':
        saldos = pd.to_numeric(mov['Saldo'], errors='coerce')
        if saldos.isna().all() or (saldos.fillna(0).abs().sum() == 0):
            advertencias.append(crear_advertencia(
                'Advertencia', archivo, banco, empresa,
                'No se detectaron saldos válidos en los movimientos.',
                'Revisar si el extracto trae la columna Saldo y si los números tienen formato correcto.'
            ))

    subset = [c for c in ['Fecha Valor', 'Descripción', 'Detalle', 'Importe', 'Saldo', 'Referencia'] if c in mov.columns]
    if subset:
        duplicados = int(mov.duplicated(subset=subset, keep=False).sum())
        if duplicados > 0:
            advertencias.append(crear_advertencia(
                'Advertencia', archivo, banco, empresa,
                f'Se detectaron {duplicados} posible(s) movimiento(s) duplicado(s).',
                'Verificar si el archivo fue exportado dos veces o si contiene movimientos repetidos.'
            ))

    cuenta = ''
    try:
        serie_cuenta = mov['Cuenta'].dropna().astype(str).map(str.strip)
        serie_cuenta = serie_cuenta[~serie_cuenta.str.lower().isin(['', 'nan', 'none'])]
        cuenta = str(serie_cuenta.iloc[0]).strip() if not serie_cuenta.empty else ''
    except Exception:
        cuenta = ''

    cuenta_invalida = (
        cuenta == ''
        or cuenta.lower() in ['nan', 'none']
        or (banco == 'PATAGONIA' and cuenta.upper() == 'PATAGONIA')
    )

    # Patagonia: evita una advertencia falsa. Algunos .XLS muestran la cuenta
    # perfectamente en Excel, pero xlrd/pandas puede devolverla vacía en la matriz.
    # Antes de avisar, se revisa el contenido binario del archivo. Si allí aparece
    # una cuenta válida tipo CC$ 385-123012668-000, se considera OK.
    if cuenta_invalida and banco == 'PATAGONIA':
        try:
            cuenta_bin, _titular_bin = extraer_cuenta_patagonia_desde_archivo(path)
            if cuenta_bin and es_codigo_cuenta_patagonia(cuenta_bin):
                cuenta_invalida = False
                if 'Cuenta' in mov.columns:
                    mov.loc[:, 'Cuenta'] = cuenta_bin
        except Exception:
            pass

    if cuenta_invalida:
        advertencias.append(crear_advertencia(
            'Advertencia', archivo, banco, empresa,
            'No se pudo leer el número/nombre de cuenta.',
            'Revisar el encabezado del extracto. Para Patagonia se espera una cuenta tipo CC$ 385-123012668-000 cerca de la etiqueta Cuenta:.'
        ))

    return advertencias


def analizar_archivo_previo(path):
    """Procesa un archivo para la vista previa, sin generar todavía el Excel final."""
    path = str(path)
    archivo = Path(path).name
    raw_preview = None

    try:
        raw_preview = read_excel_any(path, header=None)
    except Exception:
        raw_preview = None

    banco, metodo_banco = detectar_banco_detalle(path, raw_preview)
    empresa, metodo_empresa = detectar_empresa_detalle(path, raw_preview)

    registro = {
        'id': None,
        'path': path,
        'archivo': archivo,
        'banco': banco,
        'empresa': empresa,
        'tipo': '',
        'bloque': '',
        'movimientos': 0,
        'fecha_desde': None,
        'fecha_hasta': None,
        'estado': 'Error',
        'observaciones': '',
        'advertencias': [],
        'mov': None,
    }

    try:
        mov = procesar_archivo(path)
        banco = str(mov['Banco'].dropna().iloc[0]) if 'Banco' in mov.columns and not mov['Banco'].dropna().empty else banco
        empresa = str(mov['Empresa'].dropna().iloc[0]) if 'Empresa' in mov.columns and not mov['Empresa'].dropna().empty else empresa
        tipo = str(mov['Tipo Cuenta'].dropna().iloc[0]) if 'Tipo Cuenta' in mov.columns and not mov['Tipo Cuenta'].dropna().empty else ''
        bloque = str(mov['Bloque'].dropna().iloc[0]) if 'Bloque' in mov.columns and not mov['Bloque'].dropna().empty else ''
        fecha_desde, fecha_hasta = obtener_rango_fechas(mov)

        advertencias = validar_movimientos(
            path=path,
            banco=banco,
            empresa=empresa,
            tipo_cuenta=tipo,
            bloque=bloque,
            mov=mov,
            metodo_banco=metodo_banco,
            metodo_empresa=metodo_empresa
        )

        errores_reales = [a for a in advertencias if a.get('Tipo') == 'Error']
        estado = 'Error' if errores_reales else ('Advertencias' if advertencias else 'OK')
        obs = '; '.join(a['Mensaje'] for a in advertencias[:2])
        if len(advertencias) > 2:
            obs += f' (+{len(advertencias)-2} más)'

        registro.update({
            'banco': banco,
            'empresa': empresa,
            'tipo': tipo,
            'bloque': bloque,
            'movimientos': int(len(mov)),
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'estado': estado,
            'observaciones': obs or 'Sin observaciones',
            'advertencias': advertencias,
            'mov': mov,
        })

    except Exception as e:
        advertencia = crear_advertencia(
            'Error', archivo, banco, empresa,
            str(e),
            'Revisar si el Excel corresponde a BPN, Galicia, Patagonia o Mercado Pago y si fue exportado completo.'
        )
        registro.update({
            'estado': 'Error',
            'observaciones': str(e),
            'advertencias': [advertencia],
            'mov': None,
        })

    return registro


def estilo_base(ws):
    ws.sheet_view.showGridLines = False
    for col, width in {'A':24,'B':16,'C':16,'D':22,'E':38,'F':42,'G':15,'H':15,'I':18,'J':18,'K':14}.items():
        ws.column_dimensions[col].width = width


def escribir_titulo(ws, empresa, fecha):
    ws.merge_cells('A1:K1')
    c=ws['A1']
    c.value=f'MOVIMIENTOS BANCARIOS – {empresa}  |  Fecha: {fecha}'
    c.font=Font(bold=True,size=14,color='FFFFFF')
    c.fill=PatternFill('solid', fgColor='1F4E78')
    c.alignment=Alignment(horizontal='center')




def icono_banco(banco_o_bloque):
    txt = norm(banco_o_bloque)
    if 'mercado' in txt or 'mp' in txt:
        return '💳'
    if 'galicia' in txt:
        return '🟧'
    if 'patagonia' in txt:
        return '🟩'
    if 'bpn' in txt:
        return '🏦'
    return '🏛️'


def escribir_panel_lateral(ws, empresa, df_emp, destinos):
    """
    Crea un panel de navegación al costado derecho de la hoja de empresa.
    destinos: dict {bloque: fila_inicio_bloque}
    """
    panel_fill = PatternFill('solid', fgColor='1F4E78')
    sub_fill = PatternFill('solid', fgColor='D9EAF7')
    white_font = Font(bold=True, color='FFFFFF')
    blue_font = Font(bold=True, color='0563C1')
    money_fmt = '#,##0.00;[Red]-#,##0.00'
    border = Border(bottom=Side(style='thin', color='B7B7B7'))

    # Anchos del panel lateral
    for col, width in {'M':4, 'N':23, 'O':13, 'P':13, 'Q':13}.items():
        ws.column_dimensions[col].width = width

    # Título panel
    ws.merge_cells('M1:Q1')
    c = ws['M1']
    c.value = f'PANEL DE NAVEGACIÓN – {empresa}'
    c.font = Font(bold=True, size=12, color='FFFFFF')
    c.fill = panel_fill
    c.alignment = Alignment(horizontal='center')

    # Métricas rápidas
    total_mov = len(df_emp)
    ingresos = df_emp.loc[df_emp['Importe'] > 0, 'Importe'].sum()
    egresos = df_emp.loc[df_emp['Importe'] < 0, 'Importe'].sum()
    neto = df_emp['Importe'].sum()
    cant_bancos = df_emp['Banco'].nunique()

    metricas = [
        ('Bancos', cant_bancos),
        ('Movimientos', total_mov),
        ('Ingresos', ingresos),
        ('Egresos', egresos),
        ('Neto', neto),
    ]

    fila = 3
    ws.merge_cells(start_row=fila, start_column=13, end_row=fila, end_column=17)
    ws.cell(fila, 13).value = 'RESUMEN RÁPIDO'
    ws.cell(fila, 13).font = white_font
    ws.cell(fila, 13).fill = panel_fill
    ws.cell(fila, 13).alignment = Alignment(horizontal='center')
    fila += 1

    for nombre, valor in metricas:
        ws.merge_cells(start_row=fila, start_column=13, end_row=fila, end_column=15)
        ws.cell(fila, 13).value = nombre
        ws.cell(fila, 13).font = Font(bold=True)
        ws.cell(fila, 16).value = valor
        ws.merge_cells(start_row=fila, start_column=16, end_row=fila, end_column=17)
        if nombre in ['Ingresos', 'Egresos', 'Neto']:
            ws.cell(fila, 16).number_format = money_fmt
        ws.cell(fila, 16).alignment = Alignment(horizontal='right')
        for col in range(13,18):
            ws.cell(fila,col).border = border
        fila += 1

    fila += 1
    ws.merge_cells(start_row=fila, start_column=13, end_row=fila, end_column=17)
    ws.cell(fila, 13).value = 'IR A CUENTA / BANCO'
    ws.cell(fila, 13).font = white_font
    ws.cell(fila, 13).fill = panel_fill
    ws.cell(fila, 13).alignment = Alignment(horizontal='center')
    fila += 1

    # Encabezados del índice
    headers = ['Icono', 'Banco / Cuenta', 'Mov.', 'Neto', 'Saldo']
    for j, h in enumerate(headers, start=13):
        c = ws.cell(fila, j, h)
        c.font = white_font
        c.fill = PatternFill('solid', fgColor='5B9BD5')
        c.alignment = Alignment(horizontal='center')
    fila += 1

    for bloque, df_bloque in df_emp.groupby('Bloque', sort=False):
        fila_destino = destinos.get(bloque, 1)
        icono = icono_banco(bloque)
        neto_bloque = df_bloque['Importe'].sum()
        saldo_bloque = obtener_ultimo_saldo_actualizado(df_bloque)

        ws.cell(fila, 13).value = icono
        ws.cell(fila, 13).alignment = Alignment(horizontal='center')
        ws.cell(fila, 13).hyperlink = f"#'{ws.title}'!A{fila_destino}"
        ws.cell(fila, 13).style = 'Hyperlink'

        ws.cell(fila, 14).value = bloque
        ws.cell(fila, 14).hyperlink = f"#'{ws.title}'!A{fila_destino}"
        ws.cell(fila, 14).style = 'Hyperlink'
        ws.cell(fila, 14).font = blue_font
        ws.cell(fila, 14).alignment = Alignment(wrap_text=True, vertical='center')

        ws.cell(fila, 15).value = len(df_bloque)
        ws.cell(fila, 15).alignment = Alignment(horizontal='center')

        ws.cell(fila, 16).value = neto_bloque
        ws.cell(fila, 16).number_format = money_fmt
        ws.cell(fila, 16).alignment = Alignment(horizontal='right')

        ws.cell(fila, 17).value = saldo_bloque
        ws.cell(fila, 17).number_format = money_fmt
        ws.cell(fila, 17).alignment = Alignment(horizontal='right')

        for col in range(13,18):
            ws.cell(fila,col).border = border
            if fila % 2 == 0:
                ws.cell(fila,col).fill = PatternFill('solid', fgColor='F3F6FA')
        fila += 1

    fila += 1
    ws.merge_cells(start_row=fila, start_column=13, end_row=fila, end_column=17)
    ws.cell(fila, 13).value = 'Tip: hacé clic en el icono o nombre para saltar directo al bloque.'
    ws.cell(fila, 13).font = Font(italic=True, color='666666')
    ws.cell(fila, 13).alignment = Alignment(wrap_text=True)

def escribir_bloque(ws, fila, nombre_bloque, df):
    azul = PatternFill('solid', fgColor='D9EAF7')
    header_fill = PatternFill('solid', fgColor='1F4E78')
    border = Border(bottom=Side(style='thin', color='B7B7B7'))
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=11)
    cell=ws.cell(fila,1)
    cell.value=icono_banco(nombre_bloque) + '  ' + nombre_bloque + '     ↑ Volver al panel'
    cell.font=Font(bold=True, size=12, color='1F4E78')
    cell.fill=azul
    cell.alignment = Alignment(horizontal='center')
    cell.hyperlink = f"#'{ws.title}'!M1"
    cell.style = 'Hyperlink'
    fila += 1
    for j, h in enumerate(COLUMNAS_SALIDA, start=1):
        c=ws.cell(fila,j,h)
        c.font=Font(bold=True,color='FFFFFF')
        c.fill=header_fill
        c.alignment=Alignment(horizontal='center')
    fila += 1
    for _, r in df.iterrows():
        for j,h in enumerate(COLUMNAS_SALIDA, start=1):
            val=r.get(h,'')
            c=ws.cell(fila,j,val)
            c.border=border
            if h in ['Importe','Saldo']:
                c.number_format='#,##0.00;[Red]-#,##0.00'
            if 'Fecha' in h:
                c.number_format='dd/mm/yyyy'
        fila += 1
    return fila + 1


def valor_excel(valor):
    """Convierte valores pandas/numpy a valores seguros para openpyxl."""
    try:
        if pd.isna(valor):
            return None
    except Exception:
        pass

    if isinstance(valor, pd.Timestamp):
        return valor.to_pydatetime()

    return valor


def nombre_hoja_seguro(nombre):
    nombre = re.sub(r'[\/*?:\[\]]', '-', str(nombre))
    return nombre[:31] if len(nombre) > 31 else nombre


def aplicar_formato_tabla(ws, money_cols=None, date_cols=None):
    money_cols = set(money_cols or [])
    date_cols = set(date_cols or [])
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(bottom=Side(style='thin', color='D9E2F3'))

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = border
            header = ws.cell(1, c.column).value
            if header in money_cols:
                c.number_format = '#,##0.00;[Red]-#,##0.00'
            if header in date_cols:
                c.number_format = 'dd/mm/yyyy'

    for col_idx in range(1, ws.max_column + 1):
        letra = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letra]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, 55))
        ws.column_dimensions[letra].width = max_len

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def escribir_dataframe_en_hoja(wb, nombre, df, money_cols=None, date_cols=None):
    ws = wb.create_sheet(nombre_hoja_seguro(nombre))
    if df is None or df.empty:
        ws.append(['Sin datos'])
        aplicar_formato_tabla(ws)
        return ws

    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([valor_excel(row.get(col)) for col in df.columns])

    aplicar_formato_tabla(ws, money_cols=money_cols, date_cols=date_cols)
    return ws


def crear_archivos_procesados_df(registros):
    filas = []
    for r in registros or []:
        filas.append({
            'Archivo': r.get('archivo', ''),
            'Ruta': r.get('path', ''),
            'Banco': r.get('banco', ''),
            'Empresa': r.get('empresa', ''),
            'Tipo Cuenta': r.get('tipo', ''),
            'Bloque': r.get('bloque', ''),
            'Movimientos': r.get('movimientos', 0),
            'Fecha Desde': r.get('fecha_desde'),
            'Fecha Hasta': r.get('fecha_hasta'),
            'Estado': r.get('estado', ''),
            'Observaciones': r.get('observaciones', ''),
        })
    return pd.DataFrame(filas)


def crear_advertencias_df(registros=None, advertencias=None):
    filas = []
    for a in advertencias or []:
        filas.append(a)
    for r in registros or []:
        for a in r.get('advertencias', []) or []:
            filas.append(a)

    columnas = ['Tipo', 'Archivo', 'Banco', 'Empresa', 'Mensaje', 'Solución sugerida']
    if not filas:
        filas.append({
            'Tipo': 'OK',
            'Archivo': '',
            'Banco': '',
            'Empresa': '',
            'Mensaje': 'No se detectaron errores ni advertencias.',
            'Solución sugerida': '',
        })
    return pd.DataFrame(filas, columns=columnas)


def crear_totales_agrupados(todos, claves):
    filas = []
    if todos is None or todos.empty:
        return pd.DataFrame()

    for valores, df_grupo in todos.groupby(claves, dropna=False, sort=False):
        if not isinstance(valores, tuple):
            valores = (valores,)
        fila = dict(zip(claves, valores))
        importes = pd.to_numeric(df_grupo['Importe'], errors='coerce').fillna(0)
        fila['Cant. Movimientos'] = int(importes.count())
        fila['Total Ingresos'] = importes[importes > 0].sum()
        fila['Total Egresos'] = importes[importes < 0].sum()
        fila['Neto'] = importes.sum()
        fila['Último Saldo'] = obtener_ultimo_saldo_actualizado(df_grupo)
        filas.append(fila)

    return pd.DataFrame(
        filas,
        columns=claves + ['Cant. Movimientos', 'Total Ingresos', 'Total Egresos', 'Neto', 'Último Saldo']
    )


def crear_totales_por_mes(todos):
    df = todos.copy()
    fechas = obtener_fecha_orden_serie(df)
    df['_MesOrden'] = fechas.dt.to_period('M').astype(str)
    df['_MesOrden'] = df['_MesOrden'].replace('NaT', 'Sin fecha')
    return crear_totales_agrupados(df, ['_MesOrden', 'Empresa', 'Banco']).rename(columns={'_MesOrden': 'Mes'})


def generar_excel(movs, salida, archivos_procesados=None, advertencias=None):
    todos = pd.concat(movs, ignore_index=True)
    todos = normalizar_fechas_movimientos(todos)
    wb = Workbook()

    # =========================
    # RESUMEN GENERAL
    # =========================
    ws0 = wb.active
    ws0.title = 'RESUMEN_GENERAL'
    ws0.sheet_view.showGridLines = False
    ws0.freeze_panes = None
    ws0.append(['Empresa','Banco','Tipo Cuenta','Bloque','Cant. Movimientos','Total Ingresos','Total Egresos','Neto','Último Saldo'])
    resumen = crear_resumen_general(todos)
    for row in resumen.values.tolist():
        ws0.append([valor_excel(v) for v in row])
    for cell in ws0[1]:
        cell.font = Font(bold=True,color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.alignment = Alignment(horizontal='center')
    for col in range(1,10):
        ws0.column_dimensions[get_column_letter(col)].width = 20
    for row in ws0.iter_rows(min_row=2, min_col=6, max_col=9):
        for c in row:
            c.number_format = '#,##0.00;[Red]-#,##0.00'
    ws0.auto_filter.ref = ws0.dimensions

    # =========================
    # HOJAS DE TOTALES
    # =========================
    # Nota: no se crean hojas internas de control como
    # ARCHIVOS_PROCESADOS ni ERRORES_Y_ADVERTENCIAS.
    # Las observaciones se muestran únicamente en la app al finalizar.

    money_cols = ['Total Ingresos', 'Total Egresos', 'Neto', 'Último Saldo']
    escribir_dataframe_en_hoja(wb, 'TOTALES_POR_MES', crear_totales_por_mes(todos), money_cols=money_cols)
    escribir_dataframe_en_hoja(wb, 'TOTALES_POR_BANCO', crear_totales_agrupados(todos, ['Banco']), money_cols=money_cols)
    escribir_dataframe_en_hoja(wb, 'TOTALES_POR_EMPRESA', crear_totales_agrupados(todos, ['Empresa']), money_cols=money_cols)

    # =========================
    # HOJAS POR EMPRESA
    # =========================
    fecha = datetime.now().strftime('%d/%m/%Y')
    for empresa, df_emp in todos.groupby('Empresa'):
        ws = wb.create_sheet(nombre_hoja_seguro(str(empresa)))
        estilo_base(ws)
        escribir_titulo(ws, empresa, fecha)
        # Tabla principal a la izquierda y panel de navegación a la derecha.
        bloques = list(df_emp.groupby('Bloque', sort=False))

        # Precalcula las filas donde empieza cada bloque para crear hipervínculos internos.
        destinos = {}
        fila_temp = 3
        for bloque, df_bloque in bloques:
            df_temp = ordenar_por_fecha_real(df_bloque, descendente=True)
            destinos[bloque] = fila_temp
            # título del bloque + encabezado + movimientos + espacio
            fila_temp += len(df_temp) + 3

        escribir_panel_lateral(ws, empresa, df_emp, destinos)

        fila = 3
        for bloque, df_bloque in bloques:
            df_bloque = ordenar_por_fecha_real(df_bloque, descendente=True)
            fila = escribir_bloque(ws, fila, bloque, df_bloque)

        # Sin filas congeladas: no se fija la primera fila.
        ws.freeze_panes = None

    wb.save(salida)
    return salida


class App:
    def __init__(self, root):
        self.root = root
        self.root.title(f'{APP_NAME} - Procesador Bancario {APP_VERSION}')
        self.archivos = []
        self.preview_data = []
        self.preview_counter = 0
        self.carpeta_salida = self._obtener_carpeta_escritorio()
        self.nombre_salida_var = tk.StringVar(value=self._nombre_salida_default())

        self.ancho = 1120
        self.alto = 810

        # Paleta v0.2.8: más clara, empresarial y con mejor contraste.
        self.color_fondo = '#EAF0F7'
        self.color_panel = '#FFFFFF'
        self.color_panel_suave = '#F6F8FC'
        self.color_texto = '#0B2545'
        self.color_secundario = '#475569'
        self.color_azul = '#1D4ED8'
        self.color_borde = '#CBD5E1'
        self.color_verde = '#16A34A'
        self.color_gris_boton = '#334155'

        self.root.configure(bg=self.color_fondo)
        self._centrar_ventana()
        self._configurar_estilos_ttk()

        panel = tk.Frame(
            root,
            bg=self.color_panel,
            highlightthickness=1,
            highlightbackground=self.color_borde
        )
        panel.pack(fill='both', expand=True, padx=24, pady=24)

        header_frame = tk.Frame(panel, bg=self.color_panel)
        header_frame.pack(fill='x', padx=28, pady=(22, 4))

        # Columna izquierda vacía para equilibrar el botón de la derecha
        tk.Frame(
            header_frame,
            bg=self.color_panel,
            width=130
        ).pack(side='left')

        header_center = tk.Frame(header_frame, bg=self.color_panel)
        header_center.pack(side='left', expand=True, fill='x')

        tk.Label(
            header_center,
            text=f'🏦 {APP_NAME}',
            font=('Segoe UI', 24, 'bold'),
            bg=self.color_panel,
            fg=self.color_texto
        ).pack()

        tk.Label(
            header_center,
            text='Procesador Bancario por Empresa',
            font=('Segoe UI', 11, 'bold'),
            bg=self.color_panel,
            fg=self.color_secundario
        ).pack(pady=(2, 4))

        tk.Label(
            header_center,
            text=f'{APP_VERSION}  |  Software propietario',
            font=('Segoe UI', 9, 'bold'),
            bg=self.color_panel,
            fg=self.color_secundario
        ).pack()

        tk.Button(
            header_frame,
            text='ℹ Acerca de',
            command=self.mostrar_acerca_de,
            width=15,
            height=1,
            bg='#EEF2F7',
            fg=self.color_texto,
            activebackground='#E2E8F0',
            activeforeground=self.color_texto,
            relief='solid',
            bd=1,
            cursor='hand2',
            font=('Segoe UI', 9, 'bold')
        ).pack(side='right', padx=(10, 0), anchor='n')

        estilo_boton = {
            'height': 2,
            'fg': 'white',
            'activeforeground': 'white',
            'relief': 'flat',
            'cursor': 'hand2',
            'font': ('Segoe UI', 11, 'bold')
        }

        barra_botones = tk.Frame(panel, bg=self.color_panel)
        barra_botones.pack(pady=(18, 12))

        tk.Button(
            barra_botones,
            text='📂  1) Agregar archivos bancarios',
            command=self.seleccionar,
            width=31,
            bg=self.color_azul,
            activebackground='#1E40AF',
            **estilo_boton
        ).grid(row=0, column=0, padx=6)

        tk.Button(
            barra_botones,
            text='🗑️  Quitar seleccionado',
            command=self.quitar_seleccionado,
            width=23,
            bg=self.color_gris_boton,
            activebackground='#475569',
            **estilo_boton
        ).grid(row=0, column=1, padx=6)

        tk.Button(
            barra_botones,
            text='🧹  Limpiar lista',
            command=self.limpiar_lista,
            width=19,
            bg=self.color_gris_boton,
            activebackground='#475569',
            **estilo_boton
        ).grid(row=0, column=2, padx=6)

        tk.Button(
            barra_botones,
            text='📊  2) Generar resumen',
            command=self.procesar,
            width=31,
            bg=self.color_verde,
            activebackground='#166534',
            **estilo_boton
        ).grid(row=0, column=3, padx=6)

        salida_frame = tk.Frame(panel, bg=self.color_panel_suave, highlightthickness=1, highlightbackground=self.color_borde)
        salida_frame.pack(fill='x', padx=36, pady=(0, 12))

        self.lbl_salida = tk.Label(
            salida_frame,
            text=f'📁 Carpeta destinada: Escritorio  |  {self._acortar_ruta(self.carpeta_salida)}',
            bg=self.color_panel_suave,
            fg=self.color_secundario,
            anchor='w',
            font=('Segoe UI', 10, 'bold')
        )
        self.lbl_salida.pack(side='left', fill='x', expand=True, padx=12, pady=9)

        tk.Button(
            salida_frame,
            text='📁 Cambiar carpeta',
            command=lambda: self.seleccionar_carpeta_salida(forzar=True),
            width=20,
            height=1,
            bg=self.color_gris_boton,
            fg='white',
            activebackground='#475569',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            font=('Segoe UI', 10, 'bold')
        ).pack(side='right', padx=10, pady=7)


        nombre_frame = tk.Frame(panel, bg=self.color_panel_suave, highlightthickness=1, highlightbackground=self.color_borde)
        nombre_frame.pack(fill='x', padx=36, pady=(0, 12))

        tk.Label(
            nombre_frame,
            text='📝 Nombre del Excel:',
            bg=self.color_panel_suave,
            fg=self.color_texto,
            anchor='w',
            font=('Segoe UI', 10, 'bold')
        ).pack(side='left', padx=(12, 8), pady=9)

        self.ent_nombre_salida = tk.Entry(
            nombre_frame,
            textvariable=self.nombre_salida_var,
            bg='#FFFFFF',
            fg=self.color_texto,
            insertbackground=self.color_texto,
            relief='solid',
            bd=1,
            font=('Segoe UI', 10)
        )
        self.ent_nombre_salida.pack(side='left', fill='x', expand=True, padx=(0, 10), pady=7, ipady=4)

        tk.Button(
            nombre_frame,
            text='♻️ Nombre automático',
            command=self.usar_nombre_automatico,
            width=22,
            height=1,
            bg=self.color_gris_boton,
            fg='white',
            activebackground='#475569',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            font=('Segoe UI', 10, 'bold')
        ).pack(side='right', padx=10, pady=7)

        self.lbl = tk.Label(
            panel,
            text='Archivos: 0  |  OK: 0  |  Advertencias: 0  |  Errores: 0  |  Movimientos: 0',
            bg=self.color_panel,
            fg=self.color_secundario,
            font=('Segoe UI', 10, 'bold')
        )
        self.lbl.pack()

        self.progress = ttk.Progressbar(
            panel,
            orient='horizontal',
            mode='determinate',
            style='App.Horizontal.TProgressbar'
        )
        self.progress.pack(fill='x', padx=36, pady=(10, 12))

        tk.Label(
            panel,
            text='Vista previa de archivos seleccionados',
            bg=self.color_panel,
            fg=self.color_texto,
            font=('Segoe UI', 11, 'bold')
        ).pack(anchor='w', padx=26, pady=(0, 6))

        tabla_frame = tk.Frame(panel, bg=self.color_panel)
        tabla_frame.pack(fill='both', expand=True, padx=24, pady=(0, 8))

        columnas = ('archivo', 'banco', 'empresa', 'movimientos', 'estado', 'observaciones')
        self.tree = ttk.Treeview(
            tabla_frame,
            columns=columnas,
            show='headings',
            height=13,
            style='Preview.Treeview'
        )

        encabezados = {
            'archivo': ('📄 Archivo', 270),
            'banco': ('🏦 Banco', 110),
            'empresa': ('🏢 Empresa', 110),
            'movimientos': ('Mov.', 70),
            'estado': ('Estado', 120),
            'observaciones': ('Observaciones', 390),
        }

        for col, (titulo, ancho) in encabezados.items():
            self.tree.heading(col, text=titulo)
            self.tree.column(col, width=ancho, minwidth=60, anchor='center' if col in ['banco','empresa','movimientos','estado'] else 'w')

        scroll_y = ttk.Scrollbar(tabla_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll_y.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scroll_y.pack(side='right', fill='y')

        self.tree.tag_configure('OK', foreground='#14532D', background='#DCFCE7')
        self.tree.tag_configure('Advertencias', foreground='#854D0E', background='#FEF3C7')
        self.tree.tag_configure('Error', foreground='#7F1D1D', background='#FEE2E2')

        self.estado = tk.Label(
            panel,
            text='Listo para agregar archivos.',
            bg=self.color_panel,
            fg='#047857',
            wraplength=930,
            justify='center',
            font=('Segoe UI', 10)
        )
        self.estado.pack(padx=20, pady=(4, 16))

    def mostrar_acerca_de(self):
        ventana = tk.Toplevel(self.root)
        ventana.title("Acerca de XlsBank")
        ventana.configure(bg=self.color_fondo)
        ventana.transient(self.root)
        ventana.grab_set()

        ancho = 520
        alto = 390
        self.root.update_idletasks()
        x = self.root.winfo_x() + max(0, (self.root.winfo_width() - ancho) // 2)
        y = self.root.winfo_y() + max(0, (self.root.winfo_height() - alto) // 2)
        ventana.geometry(f'{ancho}x{alto}+{x}+{y}')
        ventana.resizable(False, False)

        contenedor = tk.Frame(
            ventana,
            bg=self.color_panel,
            highlightthickness=1,
            highlightbackground=self.color_borde
        )
        contenedor.pack(fill='both', expand=True, padx=18, pady=18)

        tk.Label(
            contenedor,
            text=APP_NAME,
            bg=self.color_panel,
            fg=self.color_texto,
            font=('Segoe UI', 24, 'bold')
        ).pack(pady=(24, 2))

        tk.Label(
            contenedor,
            text='Procesador Bancario',
            bg=self.color_panel,
            fg=self.color_secundario,
            font=('Segoe UI', 12, 'bold')
        ).pack(pady=(0, 14))

        info = (
            f'Versión: {APP_VERSION}\n\n'
            f'Autor: {APP_AUTHOR}\n'
            f'{APP_COPYRIGHT}\n\n'
            'Software propietario.\n'
            'Uso autorizado únicamente para clientes, equipos o usuarios habilitados.\n\n'
            'Ver LICENSE.txt para condiciones completas.'
        )

        tk.Label(
            contenedor,
            text=info,
            justify='center',
            bg=self.color_panel,
            fg=self.color_secundario,
            wraplength=430,
            font=('Segoe UI', 10)
        ).pack(padx=24, pady=(0, 18))

        tk.Button(
            contenedor,
            text='Cerrar',
            command=ventana.destroy,
            width=16,
            height=2,
            bg=self.color_azul,
            fg='white',
            activebackground='#1E40AF',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            font=('Segoe UI', 10, 'bold')
        ).pack(pady=(0, 18))    

    def _configurar_estilos_ttk(self):
        estilo = ttk.Style(self.root)
        try:
            estilo.theme_use('clam')
        except Exception:
            pass

        estilo.configure(
            'Preview.Treeview',
            background='#FFFFFF',
            fieldbackground='#FFFFFF',
            foreground='#0F172A',
            rowheight=34,
            bordercolor='#CBD5E1',
            borderwidth=1,
            font=('Segoe UI', 10)
        )
        estilo.configure(
            'Preview.Treeview.Heading',
            background='#0B2545',
            foreground='#FFFFFF',
            bordercolor='#CBD5E1',
            relief='flat',
            font=('Segoe UI', 10, 'bold')
        )
        estilo.map(
            'Preview.Treeview',
            background=[('selected', '#BFDBFE')],
            foreground=[('selected', '#0B2545')]
        )

        estilo.configure(
            'App.Horizontal.TProgressbar',
            troughcolor='#D9E2EF',
            background=self.color_azul,
            bordercolor='#D9E2EF',
            lightcolor=self.color_azul,
            darkcolor=self.color_azul,
            thickness=13
        )

    def _centrar_ventana(self):
        self.root.update_idletasks()
        ancho_pantalla = self.root.winfo_screenwidth()
        alto_pantalla = self.root.winfo_screenheight()
        x = (ancho_pantalla - self.ancho) // 2
        y = (alto_pantalla - self.alto) // 2
        self.root.geometry(
            f'{self.ancho}x{self.alto}+{x}+{y}'
        )
        self.root.minsize(self.ancho, self.alto)


    def _obtener_carpeta_escritorio(self):
        """Devuelve una carpeta de salida inicial segura: Escritorio si existe."""
        candidatos = []

        userprofile = os.environ.get('USERPROFILE')
        onedrive = os.environ.get('OneDrive') or os.environ.get('ONEDRIVE')
        home = Path.home()

        if userprofile:
            base = Path(userprofile)
            candidatos.extend([
                base / 'Desktop',
                base / 'Escritorio',
            ])

        if onedrive:
            base = Path(onedrive)
            candidatos.extend([
                base / 'Desktop',
                base / 'Escritorio',
            ])

        candidatos.extend([
            home / 'Desktop',
            home / 'Escritorio',
            home,
        ])

        for carpeta in candidatos:
            try:
                if carpeta.exists() and carpeta.is_dir():
                    return str(carpeta)
            except Exception:
                pass

        return str(home)


    def _nombre_salida_default(self):
        """Nombre automático sugerido para el Excel final."""
        return f'XlsBank_Resumen_Bancario_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

    def usar_nombre_automatico(self):
        """Regenera el nombre automático sin cambiar la carpeta elegida."""
        self.nombre_salida_var.set(self._nombre_salida_default())
        self.estado.config(
            text='Nombre automático actualizado. Podés editarlo antes de generar el Excel.',
            fg='#047857'
        )

    def _limpiar_nombre_archivo(self, nombre):
        """Limpia caracteres inválidos de Windows y asegura extensión .xlsx."""
        nombre = str(nombre or '').strip()
        if not nombre:
            nombre = self._nombre_salida_default()

        # Evita que el usuario escriba una ruta completa o caracteres inválidos.
        nombre = re.sub(r'[<>:"/\\|?*]', '-', nombre)
        nombre = re.sub(r'\s+', ' ', nombre).strip().strip('.')

        if not nombre:
            nombre = self._nombre_salida_default()

        # El programa siempre genera .xlsx.
        lower = nombre.lower()
        if lower.endswith(('.xls', '.xlsm', '.csv')):
            nombre = str(Path(nombre).with_suffix('.xlsx'))
        elif not lower.endswith('.xlsx'):
            nombre += '.xlsx'

        # Límite prudente para Windows y nombres muy largos.
        if len(nombre) > 180:
            base = Path(nombre).stem[:170]
            nombre = base + '.xlsx'

        return nombre

    def obtener_nombre_salida(self):
        """Obtiene y valida el nombre editable del Excel final."""
        nombre = self._limpiar_nombre_archivo(self.nombre_salida_var.get())
        self.nombre_salida_var.set(nombre)
        return nombre

    def _acortar_ruta(self, ruta, max_chars=78):
        ruta = str(ruta or '')
        if len(ruta) <= max_chars:
            return ruta
        return '...' + ruta[-max_chars:]

    def seleccionar_carpeta_salida(self, forzar=False):
        """Selecciona y recuerda la carpeta de salida durante la sesión."""
        if self.carpeta_salida and not forzar and Path(self.carpeta_salida).exists():
            return self.carpeta_salida

        carpeta_inicial = self.carpeta_salida if self.carpeta_salida and Path(self.carpeta_salida).exists() else None
        opciones = {'title': 'Seleccionar carpeta de salida'}
        if carpeta_inicial:
            opciones['initialdir'] = carpeta_inicial

        carpeta = filedialog.askdirectory(**opciones)
        if not carpeta:
            return None

        self.carpeta_salida = carpeta
        self.lbl_salida.config(
            text=f'📁 Carpeta destinada: {self._acortar_ruta(carpeta)}',
            fg=self.color_texto
        )
        self.estado.config(
            text='Carpeta de destino actualizada. Los próximos Excel se van a guardar ahí.',
            fg='#047857'
        )
        return carpeta

    def limpiar_lista(self):
        self.archivos = []
        self.preview_data = []
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.progress['value'] = 0
        self.lbl.config(text='Archivos seleccionados: 0')
        self.estado.config(text='Lista limpia.', fg='#047857')

    def quitar_seleccionado(self):
        seleccionados = list(self.tree.selection())
        if not seleccionados:
            messagebox.showinfo('Quitar archivo', 'Seleccioná una fila de la vista previa para quitarla.')
            return

        ids = set(seleccionados)
        self.preview_data = [r for r in self.preview_data if str(r.get('id')) not in ids]
        self.archivos = [r['path'] for r in self.preview_data]

        for item in seleccionados:
            self.tree.delete(item)

        self._actualizar_resumen_previo()

    def seleccionar(self):
        archivos = filedialog.askopenfilenames(
            title='Agregar Excel bancarios',
            filetypes=[
                ('Archivos bancarios', '*.xls *.xlsx *.csv'),
                ('Excel', '*.xls *.xlsx'),
                ('CSV', '*.csv')
            ]
        )

        if not archivos:
            return

        existentes = {str(Path(r.get('path', '')).resolve()).lower() for r in self.preview_data if r.get('path')}
        nuevos = []
        duplicados = 0

        for archivo in archivos:
            clave = str(Path(archivo).resolve()).lower()
            if clave in existentes:
                duplicados += 1
                continue
            nuevos.append(archivo)
            existentes.add(clave)

        if not nuevos:
            self._actualizar_resumen_previo()
            self.estado.config(
                text='Los archivos seleccionados ya estaban en la lista. No se agregó nada nuevo.',
                fg='#B45309'
            )
            return

        self.progress['maximum'] = max(1, len(nuevos))
        self.progress['value'] = 0
        self.estado.config(text='Analizando archivos nuevos para la vista previa...', fg='#1D4ED8')
        self.root.update()

        for idx, archivo in enumerate(nuevos, start=1):
            try:
                registro = analizar_archivo_previo(archivo)
            except Exception as e:
                nombre = Path(archivo).name
                registro = {
                    'id': None,
                    'path': archivo,
                    'archivo': nombre,
                    'banco': 'BANCO',
                    'empresa': '-',
                    'tipo': '',
                    'bloque': '',
                    'movimientos': 0,
                    'fecha_desde': None,
                    'fecha_hasta': None,
                    'estado': 'Error',
                    'observaciones': str(e),
                    'advertencias': [crear_advertencia(
                        'Error', nombre, 'BANCO', '-', str(e),
                        'Revisar si el Excel fue exportado completo y si corresponde a un banco soportado.'
                    )],
                    'mov': None,
                }

            registro['id'] = self.preview_counter
            self.preview_counter += 1
            self.preview_data.append(registro)
            self.archivos.append(archivo)
            self._insertar_registro_tabla(registro)

            self.progress['value'] = idx
            self._actualizar_resumen_previo(actualizar_estado=False)
            self.root.update()

        self._actualizar_resumen_previo()
        if duplicados:
            self.estado.config(
                text=f'Vista previa actualizada. Se agregaron {len(nuevos)} archivo(s) nuevo(s) y se omitieron {duplicados} repetido(s).',
                fg='#1D4ED8'
            )

    def _insertar_registro_tabla(self, registro):
        estado = registro.get('estado', 'Error')
        self.tree.insert(
            '',
            'end',
            iid=str(registro.get('id')),
            values=(
                registro.get('archivo', ''),
                registro.get('banco', ''),
                registro.get('empresa', ''),
                registro.get('movimientos', 0),
                estado,
                registro.get('observaciones', ''),
            ),
            tags=(estado,)
        )

    def _actualizar_resumen_previo(self, actualizar_estado=True):
        total = len(self.preview_data)
        ok = sum(1 for r in self.preview_data if r.get('estado') == 'OK')
        adv = sum(1 for r in self.preview_data if r.get('estado') == 'Advertencias')
        err = sum(1 for r in self.preview_data if r.get('estado') == 'Error')
        movs = sum(int(r.get('movimientos') or 0) for r in self.preview_data)

        self.lbl.config(
            text=f'Archivos: {total}  |  OK: {ok}  |  Advertencias: {adv}  |  Errores: {err}  |  Movimientos: {movs}'
        )

        if actualizar_estado:
            if total == 0:
                self.estado.config(text='Listo para agregar archivos.', fg='#047857')
            elif err:
                self.estado.config(
                    text='Vista previa lista. Hay archivos con error: se omitirán al generar el Excel. El detalle se mostrará en el informe final de la app.',
                    fg='#DC2626'
                )
            elif adv:
                self.estado.config(
                    text='Vista previa lista. Hay advertencias: se puede generar el Excel. El detalle se mostrará en el informe final de la app.',
                    fg='#B45309'
                )
            else:
                self.estado.config(text='Vista previa lista. Todos los archivos están OK.', fg='#047857')


    def _crear_informe_final(self, salida, registros_validos, registros_con_error):
        """
        Arma un informe simple para mostrar en pantalla al finalizar.
        El detalle por archivo solo se muestra cuando hay advertencias o errores.
        """
        total_archivos = len(self.preview_data)
        total_validos = len(registros_validos)
        total_errores = len(registros_con_error)
        total_movs = sum(int(r.get('movimientos') or 0) for r in registros_validos)

        advertencias = []
        errores = []
        for registro in self.preview_data:
            for aviso in registro.get('advertencias', []) or []:
                if aviso.get('Tipo') == 'Error':
                    errores.append(aviso)
                else:
                    advertencias.append(aviso)

        hay_observaciones = bool(advertencias or errores)

        if errores:
            estado_general = 'Finalizado con errores / archivos omitidos'
        elif advertencias:
            estado_general = 'Finalizado con advertencias para revisar'
        else:
            estado_general = 'Finalizado correctamente, sin advertencias'

        lineas = []
        lineas.append('INFORME FINAL DEL PROCESO')
        lineas.append('=' * 72)
        lineas.append('')
        lineas.append(f'Estado general: {estado_general}')
        lineas.append('')
        lineas.append('Excel generado:')
        lineas.append(str(salida))
        lineas.append('')
        lineas.append('Resumen:')
        lineas.append(f'- Archivos seleccionados: {total_archivos}')
        lineas.append(f'- Archivos procesados: {total_validos}')
        lineas.append(f'- Archivos omitidos por error: {total_errores}')
        lineas.append(f'- Movimientos procesados: {total_movs}')
        lineas.append(f'- Advertencias para revisar: {len(advertencias)}')
        lineas.append(f'- Errores detectados: {len(errores)}')
        lineas.append('')

        if not hay_observaciones:
            lineas.append('Resultado: OK')
            lineas.append('El Excel se generó correctamente y no hay observaciones para revisar.')
            lineas.append('No se muestra detalle por archivo porque todos los archivos quedaron OK.')
        else:
            lineas.append('Detalle de errores / advertencias:')
            lineas.append('-' * 72)

            for registro in self.preview_data:
                avisos = registro.get('advertencias', []) or []
                estado = registro.get('estado', '')

                # Si el archivo está OK y no tiene avisos, no se detalla.
                if estado == 'OK' and not avisos:
                    continue

                if estado == 'Error':
                    icono = 'ERROR'
                elif estado == 'Advertencias':
                    icono = 'ADVERTENCIA'
                else:
                    icono = 'OBSERVACIÓN'

                lineas.append(
                    f'{icono} | {registro.get("archivo", "")} | '
                    f'{registro.get("banco", "")} | {registro.get("empresa", "")} | '
                    f'Movimientos: {registro.get("movimientos", 0)}'
                )

                if not avisos:
                    lineas.append('  - Revisar este archivo antes de entregar el resumen.')
                else:
                    for idx, aviso in enumerate(avisos, start=1):
                        tipo = aviso.get('Tipo', 'Observación')
                        mensaje = aviso.get('Mensaje', '')
                        solucion = aviso.get('Solución sugerida', '')
                        lineas.append(f'  {idx}) Tipo: {tipo}')
                        lineas.append(f'     Qué pasó: {mensaje}')
                        if tipo == 'Error':
                            lineas.append('     Qué significa: este archivo no se usó para los movimientos del resumen final.')
                        else:
                            lineas.append('     Qué significa: la app pudo procesar el archivo, pero conviene revisar este punto.')
                        if solucion:
                            lineas.append(f'     Qué hacer: {solucion}')
                lineas.append('')

        lineas.append('')
        lineas.append('Nota: este informe se muestra solo en pantalla. No se guarda como TXT ni JSON.')
        return '\n'.join(lineas)

    def _copiar_al_portapapeles(self, texto, boton=None):
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(texto)
            self.root.update()
            if boton is not None:
                boton.config(text='Copiado')
                boton.after(1400, lambda: boton.config(text='Copiar informe'))
        except Exception:
            messagebox.showwarning('Copiar informe', 'No se pudo copiar el informe al portapapeles.')

    def _abrir_excel_generado(self, salida):
        try:
            os.startfile(salida)
        except Exception:
            messagebox.showwarning('Abrir Excel', 'No se pudo abrir el Excel automáticamente. Abrilo manualmente desde la carpeta de salida.')

    def _insertar_informe_coloreado(self, texto_widget, informe):
        """Inserta el informe aplicando color por estado: OK, advertencia o error."""
        texto_widget.tag_configure('titulo', foreground='#0B2545', font=('Consolas', 10, 'bold'))
        texto_widget.tag_configure('ok', foreground='#16A34A', font=('Consolas', 10, 'bold'))
        texto_widget.tag_configure('advertencia', foreground='#CA8A04', font=('Consolas', 10, 'bold'))
        texto_widget.tag_configure('error', foreground='#DC2626', font=('Consolas', 10, 'bold'))
        texto_widget.tag_configure('ruta', foreground='#1D4ED8')
        texto_widget.tag_configure('normal', foreground='#0F172A')

        for linea in informe.splitlines():
            upper = linea.upper()
            tag = 'normal'
            if linea.startswith('INFORME') or linea.startswith('Resumen') or linea.startswith('Detalle'):
                tag = 'titulo'
            if 'ERROR' in upper or 'OMITIDO' in upper or 'OMITIDOS' in upper:
                tag = 'error'
            elif 'ADVERTENCIA' in upper or 'ADVERTENCIAS' in upper:
                tag = 'advertencia'
            elif 'RESULTADO: OK' in upper or 'CORRECTAMENTE' in upper or 'SIN ADVERTENCIAS' in upper:
                tag = 'ok'
            elif '.XLSX' in upper or '.XLS' in upper:
                tag = 'ruta'

            texto_widget.insert('end', linea + '\n', tag)

    def _mostrar_informe_final(self, salida, registros_validos, registros_con_error):
        """Muestra una ventana clara con el resultado del proceso, sin crear archivos extra."""
        informe = self._crear_informe_final(salida, registros_validos, registros_con_error)

        ventana = tk.Toplevel(self.root)
        ventana.title('Informe del proceso')
        ventana.configure(bg=self.color_fondo)
        ventana.transient(self.root)
        ventana.grab_set()

        tiene_error = any(r.get('estado') == 'Error' for r in self.preview_data)
        tiene_adv = any(r.get('estado') == 'Advertencias' for r in self.preview_data)
        hay_observaciones = tiene_error or tiene_adv

        ancho = 900
        alto = 640 if hay_observaciones else 500
        self.root.update_idletasks()
        x = self.root.winfo_x() + max(0, (self.root.winfo_width() - ancho) // 2)
        y = self.root.winfo_y() + max(0, (self.root.winfo_height() - alto) // 2)
        ventana.geometry(f'{ancho}x{alto}+{x}+{y}')
        ventana.minsize(760, 470)

        contenedor = tk.Frame(
            ventana,
            bg=self.color_panel,
            highlightthickness=1,
            highlightbackground=self.color_borde
        )
        contenedor.pack(fill='both', expand=True, padx=18, pady=18)

        if tiene_error:
            titulo = '🔴 Proceso finalizado con errores'
            subtitulo = 'Hay archivos omitidos. Revisá el detalle antes de usar el Excel generado.'
            color_titulo = '#DC2626'
            color_fondo_estado = '#FEE2E2'
            color_borde_estado = '#FCA5A5'
        elif tiene_adv:
            titulo = '🟡 Proceso finalizado con advertencias'
            subtitulo = 'El Excel se generó, pero hay puntos para revisar.'
            color_titulo = '#CA8A04'
            color_fondo_estado = '#FEF3C7'
            color_borde_estado = '#FACC15'
        else:
            titulo = '✅ Proceso finalizado correctamente'
            subtitulo = 'El Excel se generó correctamente. No hay advertencias ni errores para revisar.'
            color_titulo = '#16A34A'
            color_fondo_estado = '#DCFCE7'
            color_borde_estado = '#86EFAC'

        estado_frame = tk.Frame(
            contenedor,
            bg=color_fondo_estado,
            highlightthickness=1,
            highlightbackground=color_borde_estado
        )
        estado_frame.pack(fill='x', padx=18, pady=(18, 12))

        tk.Label(
            estado_frame,
            text=titulo,
            bg=color_fondo_estado,
            fg=color_titulo,
            font=('Segoe UI', 18, 'bold')
        ).pack(pady=(14, 4))

        tk.Label(
            estado_frame,
            text=subtitulo,
            bg=color_fondo_estado,
            fg='#334155',
            font=('Segoe UI', 10, 'bold')
        ).pack(pady=(0, 14))

        marco_texto = tk.Frame(contenedor, bg=self.color_panel)
        marco_texto.pack(fill='both', expand=True, padx=18, pady=(0, 12))

        scroll = ttk.Scrollbar(marco_texto, orient='vertical')
        texto = tk.Text(
            marco_texto,
            wrap='word',
            yscrollcommand=scroll.set,
            bg='#FFFFFF',
            fg='#0F172A',
            insertbackground='#0F172A',
            relief='solid',
            bd=1,
            padx=14,
            pady=12,
            font=('Consolas', 10)
        )
        scroll.config(command=texto.yview)
        texto.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')
        self._insertar_informe_coloreado(texto, informe)
        texto.config(state='disabled')

        barra = tk.Frame(contenedor, bg=self.color_panel)
        barra.pack(fill='x', padx=18, pady=(0, 16))

        btn_copiar = tk.Button(
            barra,
            text='📋 Copiar informe',
            command=lambda: self._copiar_al_portapapeles(informe, btn_copiar),
            width=18,
            height=2,
            bg=self.color_gris_boton,
            fg='white',
            activebackground='#475569',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            font=('Segoe UI', 9, 'bold')
        )
        btn_copiar.pack(side='left')

        tk.Button(
            barra,
            text='📂 Abrir Excel generado',
            command=lambda: self._abrir_excel_generado(salida),
            width=24,
            height=2,
            bg=self.color_verde,
            fg='white',
            activebackground='#15803D',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            font=('Segoe UI', 9, 'bold')
        ).pack(side='right', padx=(8, 0))

        tk.Button(
            barra,
            text='Cerrar',
            command=ventana.destroy,
            width=14,
            height=2,
            bg=self.color_azul,
            fg='white',
            activebackground='#1F63D3',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            font=('Segoe UI', 9, 'bold')
        ).pack(side='right')

    def procesar(self):
        if not self.preview_data:
            messagebox.showerror(
                'Faltan archivos',
                'Primero seleccioná uno o varios Excel para generar la vista previa.'
            )
            return

        registros_validos = [r for r in self.preview_data if r.get('mov') is not None and r.get('estado') != 'Error']
        registros_con_error = [r for r in self.preview_data if r.get('estado') == 'Error']

        if not registros_validos:
            messagebox.showerror(
                'No hay archivos válidos',
                'Todos los archivos quedaron con error. Revisá la vista previa antes de generar el resumen.'
            )
            return

        if registros_con_error:
            continuar = messagebox.askyesno(
                'Hay archivos con error',
                'Hay uno o más archivos con error. Se van a omitir en los movimientos, pero quedarán informados en el informe final de la app.\n\n¿Querés continuar?'
            )
            if not continuar:
                return

        carpeta = self.seleccionar_carpeta_salida(forzar=False)

        if not carpeta:
            return

        try:
            self.estado.config(text='Generando Excel final...', fg='#1D4ED8')
            self.root.update()

            movs = [r['mov'] for r in registros_validos]

            nombre = self.obtener_nombre_salida()
            if not nombre:
                return

            salida_path = Path(carpeta) / nombre
            if salida_path.exists():
                reemplazar = messagebox.askyesno(
                    'El archivo ya existe',
                    f'Ya existe un Excel con este nombre:\n\n{salida_path}\n\n¿Querés reemplazarlo?'
                )
                if not reemplazar:
                    self.estado.config(
                        text='Proceso cancelado. Cambiá el nombre del Excel o elegí reemplazar el archivo existente.',
                        fg='#92400E'
                    )
                    return

            salida = str(salida_path)

            generar_excel(
                movs=movs,
                salida=salida,
                archivos_procesados=self.preview_data,
                advertencias=None
            )

            total_movs = sum(int(r.get('movimientos') or 0) for r in registros_validos)
            msg = (
                f'Listo. Archivo generado:\n{salida}\n\n'
                f'Archivos procesados: {len(registros_validos)}\n'
                f'Movimientos procesados: {total_movs}'
            )

            if registros_con_error:
                msg += f'\nArchivos omitidos por error: {len(registros_con_error)}'

            self.estado.config(text=msg, fg='#047857')
            self._mostrar_informe_final(salida, registros_validos, registros_con_error)

        except Exception as e:
            self.estado.config(text='Error al procesar.', fg='#DC2626')
            messagebox.showerror('Error', str(e))

def iniciar_aplicacion(root):
    root.app = App(root)
    root.deiconify()
    root.lift()
    root.focus_force()


if __name__ == '__main__':
    root = tk.Tk()
    root.withdraw()

    ruta_gif = ruta_recurso(
        Path('assets') / 'logo_girando_desde_cero_sin_fondo.gif'
    )

    try:
        if ruta_gif.exists():
            PantallaCarga(
                root=root,
                ruta_gif=ruta_gif,
                al_terminar=lambda: iniciar_aplicacion(root),
                ancho=760,
                alto=720
            )
        else:
            iniciar_aplicacion(root)
            messagebox.showwarning(
                'Pantalla de carga',
                f'No se encontró el GIF:\n{ruta_gif}'
            )

    except Exception as error:
        # Si falla el splash, no dejamos una ventana negra encima.
        for ventana in root.winfo_children():
            try:
                if isinstance(ventana, tk.Toplevel):
                    ventana.destroy()
            except Exception:
                pass

        iniciar_aplicacion(root)
        messagebox.showwarning(
            'Pantalla de carga',
            f'No se pudo mostrar la animación:\n{error}'
        )

    root.mainloop()